'''
import importlib
import aidulib
importlib.reload(aidulib)
'''

# AIDU Library Import
from aicentro.session import Session
from aicentro.framework.keras import Keras as AiduFrm
aidu_session = Session(verify=False)
aidu_framework = AiduFrm(session=aidu_session)


# 'path'내 존재하는 파일 읽기.
import os
def get_flst(path='', option='t', show=False):
    # path: 타깃경로
    # option: t=해당폴더만, r=하위경로 포함
    # show: 읽어낸 경로 및 파일명 시각화
    flst = []
    path = aidu_framework.config.data_dir+'/'+path
    
    if option == 't':  # 지정 폴더 탐색
        lst = os.listdir(path)
        for dir in lst:
            flst.append(path+'/'+dir)
            
    if option == 'r':  # 하위 폴더 전체 탐색
        for root, dirs, files in os.walk(path):
            for file in files:
                file_path = os.path.join(root, file)
                flst.append(file_path)
    
    if flst:
        if show:
            for fname in flst:
                print(fname)
    else:
        print('File not found.')
    return flst



get_ipython().system('pip3 install openpyxl')
import pandas as pd
from openpyxl import load_workbook

# xlsx -> csv
def xlsx_to_csv(in_path, out_path):
    # inpath: 파일명 (get_flst 사용해서 경로 입력)
    # out_path: 저장 경로 (폴더명 지정시 자동 폴더 생성 및 저장)
    if not in_path.endswith(".xlsx"):
        return
    os.makedirs(out_path, exist_ok=True)  # 저장 경로 생성
    
    wb = load_workbook(in_path, read_only=True, keep_links=False)
    sheets = wb.sheetnames
    out_file = in_path.split('/')[-1].split('.xlsx')[0]
        
    if len(sheets) > 1:
        print("Multiple Sheets detected. Generating Multi-CSV...")
        for fname in sheets:
            xlsx = pd.read_excel(in_path, sheet_name=fname ,engine='openpyxl')
            xlsx.to_csv(f'{out_path}/{out_file}_{fname}.csv', index=False)
            print(f'{fname} Done!')
    else:
        xlsx = pd.read_excel(in_path, engine='openpyxl')
        xlsx.to_csv(f'{out_path}/{out_file}.csv', index=False)
        print(f'{out_file} Done!')
        
        
import os
from tqdm import tqdm

# path -> 데이터관리/download
# 주의) 실행시 저장할 경로(download)내 모든 파일을 지운다!
def readyDL(path): 
    # path: 옮길 파일들이 있는 경로
    save_path = aidu_framework.config.data_dir+'/download/' # 저장할 곳
    from_path =  os.path.join(os.getcwd(), path)
    print(f'Start moving {path} to download.\nCleaning download folder...')
    
    os.system(f'rm -r -f {save_path}')  # download 삭제
    os.makedirs(save_path, exist_ok=True)  # 재생성
    
    flst = os.listdir(path)  # 옮길 목록

    for file in tqdm(flst, desc='Copying...'):
        indata = os.path.join(from_path, file)
        outdata = os.path.join(save_path, file)
        file_path, file_name = os.path.split(outdata)
        os.system(f'cp {indata} {outdata}')
    print(f'Done!')
    
    
import shutil
# 지정 경로 이하의 모든 폴더 + 파일 싹지움
def delall(path):  # yes/no 안물어보니 신중하게 사용!
    shutil.rmtree(path)

# 빈파일 만들기 1년치 데이터를 step 간격으로 뽑아낸다. 윤년계산은 없으니 주의
def genEmpty(year, step):
    ans = pd.DataFrame(columns=['Timestamp'])
    tmp_lst = []
    for month in tqdm(range(1, 13)):
        for day in range(1, 32):
            if month == 2 and day > 28:
                continue
            if month in [4, 6, 9, 11] and day > 30:
                continue
            for hour in range(24):
                for mn in range(0,60,step):
                    date = str(year)+str(month).zfill(2) + str(day).zfill(2)
                    time = str(hour).zfill(2) + str(mn).zfill(2)
                    nxt = str((hour + 1)%24 if mn+step == 60 else hour).zfill(2) + str((mn+step)%60).zfill(2)
                    stn = date + '_' + time + '-' + nxt
                    tmp_lst.append(stn)
    result_df = pd.DataFrame(tmp_lst, columns=['Timestamp'])            
    return result_df



from collections import Counter
# dataframe 내 특정 컬럼의 유니크값 추출
def count_unique(df, col, print_count=False):
    print(f"중복되지 않는 {col} 종류:", df[f"{col}"].unique())
    print(f"중복되지 않는 {col} 데이터 수:", df[f"{col}"].nunique())
    if print_count:
        freq = Counter(df[col])
        sorted_freq = sorted(freq.items(), key = lambda item: item[1], reverse=True)
        for k, v in sorted_freq:
            print(f'{k}: {v}')
        
# dataframe내 특정 컬럼에서 특정한 값이 있는지 확인
def check_val(df, col, val):
    print(df[df[col] == val])

# dataframe 내 결측치 있는 샘플 추출
def find_missing(df, col=None):
    missing = df.isnull().sum()
    print(missing)
    if col:
        missing_df = merged_df[merged_df[col].isnull()]
        missing_df.to_csv('missing.csv', index=False)

        
# 중복값 추출
def find_dup(df, col):
    dups = df[df.duplicated(col, keep=False)]
    return dups

# 두 데이터프레임의 특정 컬럼을 기준으로 동일한 값을 오른쪽으로 밀어내기
def dupsToRight(left, right, col):
    left_key = left[col].unique()
    right_key = right[col].unique()
    target = list(set(left_key).intersection(right_key))
    
    if target:
        print(f"Pushing DUPS in [{col}]: \n{target}")
        dups = left[left[col].isin(target)]
        left = left[~left[col].isin(target)]
        right = pd.concat([right, dups])
    else:
        print("No Dups found.")

    return left, right



import seaborn as sns
import matplotlib.pyplot as plt
from sklearn.metrics import confusion_matrix, classification_report, f1_score

def show_cm(label, pred, img=False):
    cm = confusion_matrix(label, pred)
    if img:
        plt.figure(figsize=(10, 7))
        sns.heatmap(cm, annot=True, fmt="d", cmap="Blues")
        plt.title("Confusion Matrix")
        plt.xlabel("predicted Labels")
        plt.ylabel("True Labels")
        plt.show()

    else:
        print(cm)
    report = classification_report(label, pred, zero_division=1)
    print(report)
    f1 = f1_score(label, pred, average='weighted')
    print(f'F1-score: {f1}')
    
    
def heatmap(df, cols=None, fname='tmp.png'):
    # 상관관계 계산
    correlation_matrix = df.corr()
    # 시각화 - 모든 컬럼을 포함한 상관관계 히트맵
    plt.figure(figsize=(12, 10))
    sns.heatmap(correlation_matrix, cmap='coolwarm', annot=False, fmt=".2f", linewidths=0.5)
    plt.title('Correlation Heatmap of all Columns')

    # 이미지로 저장
    plt.savefig(fname)
    plt.show()
    plt.close()
    
    if cols:
        for col in cols:
        # 'uenoavg', 'uenomax'와의 상관관계를 기준으로 내림차순으로 정렬
            sorted_corr = correlation_matrix[col].sort_values(ascending=False)

        # 연관성이 높은 상위 20개 컬럼명 출력
        top_related = sorted_corr[1:]  # 첫 번째는 자기 자신이므로 제외

        print("Top columns correlated with uenoavg:")
        print(top_related)


    